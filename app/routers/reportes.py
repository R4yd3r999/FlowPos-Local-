from __future__ import annotations

import io
import sqlite3

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.deps import get_conn
from app.errors import NotFoundError
from app.roles import PUEDE_OPERAR_CAJA, PUEDE_VER_INSUMOS
from app.routers.auth import require_role
from app.services import caja as caja_service
from app.services import conteo as conteo_service
from app.services import inventario

router = APIRouter(prefix="/reportes", tags=["reportes"])

FUENTE = "Arial"
FORMATO_CUP = '#,##0.00 "CUP"'

TIPO_MOVIMIENTO_LABEL = {
    "entrada_compra": "Entrada por compra",
    "salida_venta": "Salida por venta",
    "reversa_cancelacion": "Reversa (cancelación)",
    "salida_consumo_interno": "Salida — consumo interno",
    "salida_merma": "Salida — merma/rotura",
    "salida_otro": "Salida — otro",
    "ajuste_conteo": "Ajuste por conteo",
}


def _hoja_con_estilo(wb: Workbook, nombre: str):
    ws = wb.create_sheet(nombre)
    return ws


RELLENO_ENCABEZADO = PatternFill(start_color="B8871E", end_color="B8871E", fill_type="solid")


def _encabezado(ws, fila: int, columnas: list[str]) -> None:
    for col_idx, texto in enumerate(columnas, start=1):
        celda = ws.cell(row=fila, column=col_idx, value=texto)
        celda.font = Font(name=FUENTE, bold=True, color="FFFFFF")
        celda.fill = RELLENO_ENCABEZADO
        celda.alignment = Alignment(horizontal="center")
    for col_idx in range(1, len(columnas) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 20


def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/ventas.xlsx")
def exportar_ventas(
    usuario_id: int,
    desde: str | None = None,
    hasta: str | None = None,
    conn: sqlite3.Connection = Depends(get_conn),
) -> StreamingResponse:
    require_role(conn, usuario_id, *PUEDE_VER_INSUMOS)
    query = """
        SELECT c.id AS cuenta_id, c.referencia, ci.id AS item_id, p.nombre AS producto,
               ci.cantidad, ci.precio_unitario_aplicado, ci.estado,
               ci.agregado_at, ci.motivo_cancelacion
        FROM cuenta_items ci
        JOIN cuentas c ON c.id = ci.cuenta_id
        JOIN productos p ON p.id = ci.producto_id
        WHERE 1 = 1
    """
    params: list = []
    if desde:
        query += " AND ci.agregado_at >= ?"
        params.append(desde)
    if hasta:
        query += " AND ci.agregado_at <= ?"
        params.append(hasta)
    query += " ORDER BY ci.agregado_at"
    filas = conn.execute(query, params).fetchall()

    pagos_query = """
        SELECT p.id, p.cuenta_id, c.referencia, p.metodo, p.moneda, p.subtipo,
               p.monto, p.monto_cup_equivalente, p.registrado_at
        FROM pagos p JOIN cuentas c ON c.id = p.cuenta_id
        WHERE 1 = 1
    """
    if desde:
        pagos_query += " AND p.registrado_at >= ?"
    if hasta:
        pagos_query += " AND p.registrado_at <= ?"
    pagos_query += " ORDER BY p.registrado_at"
    pagos = conn.execute(pagos_query, params).fetchall()

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = _hoja_con_estilo(wb, "Ventas")
    _encabezado(ws1, 1, ["Cuenta", "Referencia", "Producto", "Cantidad", "Precio unit. (CUP)", "Subtotal (CUP)", "Estado", "Fecha", "Motivo cancelación"])
    fila = 2
    total_confirmado = 0
    for r in filas:
        subtotal_cup = (r["precio_unitario_aplicado"] * r["cantidad"]) / 100
        ws1.cell(row=fila, column=1, value=r["cuenta_id"])
        ws1.cell(row=fila, column=2, value=r["referencia"])
        ws1.cell(row=fila, column=3, value=r["producto"])
        ws1.cell(row=fila, column=4, value=r["cantidad"])
        ws1.cell(row=fila, column=5, value=r["precio_unitario_aplicado"] / 100).number_format = FORMATO_CUP
        ws1.cell(row=fila, column=6, value=subtotal_cup).number_format = FORMATO_CUP
        ws1.cell(row=fila, column=7, value=r["estado"])
        ws1.cell(row=fila, column=8, value=r["agregado_at"])
        ws1.cell(row=fila, column=9, value=r["motivo_cancelacion"] or "")
        if r["estado"] == "confirmado":
            total_confirmado += subtotal_cup
        fila += 1
    ws1.cell(row=fila + 1, column=5, value="Total confirmado:").font = Font(name=FUENTE, bold=True)
    ws1.cell(row=fila + 1, column=6, value=total_confirmado).number_format = FORMATO_CUP

    ws2 = _hoja_con_estilo(wb, "Pagos")
    _encabezado(ws2, 1, ["ID pago", "Cuenta", "Referencia", "Método", "Moneda", "Subtipo", "Monto", "Equivalente CUP", "Fecha"])
    fila = 2
    for r in pagos:
        ws2.cell(row=fila, column=1, value=r["id"])
        ws2.cell(row=fila, column=2, value=r["cuenta_id"])
        ws2.cell(row=fila, column=3, value=r["referencia"])
        ws2.cell(row=fila, column=4, value=r["metodo"])
        ws2.cell(row=fila, column=5, value=r["moneda"])
        ws2.cell(row=fila, column=6, value=r["subtipo"] or "")
        ws2.cell(row=fila, column=7, value=r["monto"] / 100)
        ws2.cell(row=fila, column=8, value=r["monto_cup_equivalente"] / 100).number_format = FORMATO_CUP
        ws2.cell(row=fila, column=9, value=r["registrado_at"])
        fila += 1

    for ws in (ws1, ws2):
        for row in ws.iter_rows():
            for celda in row:
                if celda.font is None or not celda.font.bold:
                    celda.font = Font(name=FUENTE)

    return _xlsx_response(wb, "ventas.xlsx")


@router.get("/inventario.xlsx")
def exportar_inventario(usuario_id: int, conn: sqlite3.Connection = Depends(get_conn)) -> StreamingResponse:
    require_role(conn, usuario_id, *PUEDE_VER_INSUMOS)
    filas = conn.execute(
        "SELECT nombre, unidad_medida, cantidad_actual, cantidad_minima, costo_promedio, activo FROM insumos ORDER BY nombre"
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    _encabezado(ws, 1, ["Insumo", "Unidad", "Cantidad actual", "Cantidad mínima", "Bajo mínimo", "Costo promedio (CUP)", "Valor total (CUP)", "Activo"])

    fila = 2
    for r in filas:
        bajo_minimo = r["cantidad_actual"] <= r["cantidad_minima"]
        valor_total = (r["cantidad_actual"] * r["costo_promedio"]) / 100
        ws.cell(row=fila, column=1, value=r["nombre"], )
        ws.cell(row=fila, column=2, value=r["unidad_medida"])
        ws.cell(row=fila, column=3, value=r["cantidad_actual"])
        ws.cell(row=fila, column=4, value=r["cantidad_minima"])
        celda_alerta = ws.cell(row=fila, column=5, value="SÍ" if bajo_minimo else "no")
        if bajo_minimo:
            celda_alerta.font = Font(name=FUENTE, bold=True, color="C0392B")
        ws.cell(row=fila, column=6, value=r["costo_promedio"] / 100).number_format = FORMATO_CUP
        ws.cell(row=fila, column=7, value=valor_total).number_format = FORMATO_CUP
        ws.cell(row=fila, column=8, value="Sí" if r["activo"] else "No")
        fila += 1

    for row in ws.iter_rows():
        for celda in row:
            if celda.font is None or not celda.font.bold:
                celda.font = Font(name=FUENTE)

    return _xlsx_response(wb, "inventario.xlsx")


def _hoja_movimientos(ws, filas: list[dict]) -> None:
    _encabezado(ws, 1, ["Fecha", "Insumo", "Tipo", "Cantidad", "Resultante", "Referencia", "Usuario", "Nota"])
    fila = 2
    for r in filas:
        referencia = r["referencia_tipo"] or ""
        if r["referencia_id"]:
            referencia += f' #{r["referencia_id"]}'
        ws.cell(row=fila, column=1, value=r["created_at"])
        ws.cell(row=fila, column=2, value=r["insumo_nombre"])
        ws.cell(row=fila, column=3, value=TIPO_MOVIMIENTO_LABEL.get(r["tipo"], r["tipo"]))
        signo = "+" if r["cantidad"] > 0 else ""
        ws.cell(row=fila, column=4, value=f'{signo}{r["cantidad"]} {r["unidad_medida"]}')
        ws.cell(row=fila, column=5, value=f'{r["cantidad_resultante"]} {r["unidad_medida"]}')
        ws.cell(row=fila, column=6, value=referencia)
        ws.cell(row=fila, column=7, value=r["usuario_nombre"])
        ws.cell(row=fila, column=8, value=r["nota"] or "")
        fila += 1
    if fila == 2:
        ws.cell(row=2, column=1, value="Sin movimientos en este período")


@router.get("/movimientos.xlsx")
def exportar_movimientos(
    usuario_id: int,
    desde: str | None = None,
    hasta: str | None = None,
    insumo_id: int | None = None,
    tipo: str | None = None,
    conn: sqlite3.Connection = Depends(get_conn),
) -> StreamingResponse:
    """Exporta el ledger completo de movimientos de inventario (todas las
    entradas y salidas, con o sin caja abierta), con los mismos filtros
    que la pestaña de pantalla."""
    require_role(conn, usuario_id, *PUEDE_VER_INSUMOS)
    filas = [
        dict(r)
        for r in inventario.listar_movimientos(
            conn, insumo_id=insumo_id, tipo=tipo, desde=desde, hasta=hasta, limit=100000, offset=0
        )
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    _hoja_movimientos(ws, filas)

    for row in ws.iter_rows():
        for celda in row:
            if celda.font is None or not celda.font.bold:
                celda.font = Font(name=FUENTE)

    return _xlsx_response(wb, "movimientos_inventario.xlsx")


@router.get("/movimientos.pdf")
def exportar_movimientos_pdf(
    usuario_id: int,
    desde: str | None = None,
    hasta: str | None = None,
    insumo_id: int | None = None,
    tipo: str | None = None,
    conn: sqlite3.Connection = Depends(get_conn),
) -> StreamingResponse:
    """Versión imprimible/archivable del mismo ledger filtrado que .xlsx exporta como datos."""
    require_role(conn, usuario_id, *PUEDE_VER_INSUMOS)
    filas = [
        dict(r)
        for r in inventario.listar_movimientos(
            conn, insumo_id=insumo_id, tipo=tipo, desde=desde, hasta=hasta, limit=100000, offset=0
        )
    ]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    elementos = [Paragraph("FlowPos (Local) — Movimientos de inventario", _ESTILO_TITULO)]
    filtros_txt = []
    if desde:
        filtros_txt.append(f"desde {desde}")
    if hasta:
        filtros_txt.append(f"hasta {hasta}")
    if insumo_id:
        nombre_insumo = filas[0]["insumo_nombre"] if filas else str(insumo_id)
        filtros_txt.append(f"insumo: {nombre_insumo}")
    if tipo:
        filtros_txt.append(f"tipo: {TIPO_MOVIMIENTO_LABEL.get(tipo, tipo)}")
    subtitulo = (" · ".join(filtros_txt)) if filtros_txt else "Todos los movimientos"
    elementos.append(Paragraph(subtitulo, _ESTILO_SUBTITULO))
    elementos.append(Spacer(1, 10))

    filas_tabla = [
        [
            m["created_at"], m["insumo_nombre"], TIPO_MOVIMIENTO_LABEL.get(m["tipo"], m["tipo"]),
            f'{"+" if m["cantidad"] > 0 else ""}{m["cantidad"]} {m["unidad_medida"]}',
            m["usuario_nombre"], m["nota"] or "",
        ]
        for m in filas
    ]
    elementos.append(_tabla_pdf(
        ["Fecha", "Insumo", "Tipo", "Cant.", "Usuario", "Nota"], filas_tabla,
        anchos=[3.2 * cm, 3.2 * cm, 3.6 * cm, 2.2 * cm, 2.5 * cm, 4 * cm],
    ))

    doc.build(elementos)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="movimientos_inventario.pdf"'},
    )


def _cierre_o_404(conn: sqlite3.Connection, cierre_id: int) -> sqlite3.Row:
    row = conn.execute("SELECT * FROM cierres_caja WHERE id = ?", (cierre_id,)).fetchone()
    if row is None:
        raise NotFoundError(f"Cierre {cierre_id} no existe")
    return row


def _datos_dia_cierre(conn: sqlite3.Connection, cierre_id: int) -> dict:
    """Reune todo lo de un día de caja -- resumen de pagos, ventas, pagos uno
    por uno, movimientos de inventario y los dos conteos -- reusando
    exactamente las mismas fuentes que ya se muestran en pantalla en
    /cierre-caja/{id}/detalle, para que lo exportado nunca pueda descuadrar
    con lo que el operador ya vio."""
    cierre = _cierre_o_404(conn, cierre_id)
    hasta = cierre["hora_cierre"]
    resumen = caja_service.resumen_periodo(conn, cierre["hora_apertura"], hasta)
    ventas = caja_service.ventas_periodo(conn, cierre["hora_apertura"], hasta)
    pagos = caja_service.pagos_detalle_periodo(conn, cierre["hora_apertura"], hasta)
    movimientos = caja_service.movimientos_periodo(conn, cierre["hora_apertura"], hasta)
    conteos = conteo_service.obtener_conteos(conn, cierre_id)

    abierto_por = conn.execute("SELECT nombre FROM usuarios WHERE id = ?", (cierre["abierto_por_id"],)).fetchone()
    cerrado_por = None
    if cierre["cerrado_por_id"]:
        cerrado_por = conn.execute(
            "SELECT nombre FROM usuarios WHERE id = ?", (cierre["cerrado_por_id"],)
        ).fetchone()

    return {
        "cierre": dict(cierre),
        "abierto_por_nombre": abierto_por["nombre"] if abierto_por else "—",
        "cerrado_por_nombre": cerrado_por["nombre"] if cerrado_por else None,
        "resumen": resumen,
        "ventas": ventas,
        "pagos": pagos,
        "movimientos": movimientos,
        "conteo_apertura": [c for c in conteos if c["momento"] == "apertura"],
        "conteo_cierre": [c for c in conteos if c["momento"] == "cierre"],
    }


@router.get("/cierre/{cierre_id}.xlsx")
def exportar_cierre_xlsx(
    cierre_id: int, usuario_id: int, conn: sqlite3.Connection = Depends(get_conn)
) -> StreamingResponse:
    """Exporta el día de caja completo en un solo archivo: pagos, ventas,
    movimientos de inventario y los dos conteos, cada uno en su hoja."""
    require_role(conn, usuario_id, *PUEDE_OPERAR_CAJA)
    d = _datos_dia_cierre(conn, cierre_id)
    cierre = d["cierre"]

    wb = Workbook()
    wb.remove(wb.active)

    ws_resumen = wb.create_sheet("Resumen")
    ws_resumen.cell(row=1, column=1, value=f"Cierre de caja — {cierre['fecha']}").font = Font(
        name=FUENTE, bold=True, size=14
    )
    ws_resumen.cell(row=2, column=1, value=f"Abrió: {d['abierto_por_nombre']} · {cierre['hora_apertura']}")
    if cierre["hora_cierre"]:
        ws_resumen.cell(
            row=3, column=1, value=f"Cerró: {d['cerrado_por_nombre'] or '—'} · {cierre['hora_cierre']}"
        )
    if cierre["diferencia_cup"] is not None:
        ws_resumen.cell(row=4, column=1, value="Efectivo contado (CUP)").font = Font(name=FUENTE, bold=True)
        ws_resumen.cell(row=4, column=2, value=cierre["efectivo_contado_cup"] / 100).number_format = FORMATO_CUP
        ws_resumen.cell(row=5, column=1, value="Efectivo esperado (CUP)").font = Font(name=FUENTE, bold=True)
        ws_resumen.cell(row=5, column=2, value=cierre["efectivo_esperado_cup"] / 100).number_format = FORMATO_CUP
        ws_resumen.cell(row=6, column=1, value="Diferencia (CUP)").font = Font(name=FUENTE, bold=True)
        ws_resumen.cell(row=6, column=2, value=cierre["diferencia_cup"] / 100).number_format = FORMATO_CUP
    _encabezado(ws_resumen, 8, ["Método", "Moneda", "Subtipo", "Pagos", "Total (CUP)"])
    fila = 9
    for f in d["resumen"]["desglose"]:
        ws_resumen.cell(row=fila, column=1, value=f["metodo"])
        ws_resumen.cell(row=fila, column=2, value=f["moneda"])
        ws_resumen.cell(row=fila, column=3, value=f["subtipo"] or "")
        ws_resumen.cell(row=fila, column=4, value=f["cantidad"])
        ws_resumen.cell(row=fila, column=5, value=f["monto_cup_total"] / 100).number_format = FORMATO_CUP
        fila += 1

    ws_ventas = wb.create_sheet("Ventas")
    _encabezado(ws_ventas, 1, ["Cuenta", "Referencia", "Producto", "Cantidad", "Precio unit. (CUP)", "Subtotal (CUP)", "Estado", "Fecha", "Motivo cancelación"])
    fila = 2
    total_confirmado = 0
    for v in d["ventas"]:
        subtotal_cup = v["subtotal"] / 100
        ws_ventas.cell(row=fila, column=1, value=v["id"])
        ws_ventas.cell(row=fila, column=2, value=v["referencia"])
        ws_ventas.cell(row=fila, column=3, value=v["producto"])
        ws_ventas.cell(row=fila, column=4, value=v["cantidad"])
        ws_ventas.cell(row=fila, column=5, value=v["precio_unitario_aplicado"] / 100).number_format = FORMATO_CUP
        ws_ventas.cell(row=fila, column=6, value=subtotal_cup).number_format = FORMATO_CUP
        ws_ventas.cell(row=fila, column=7, value=v["estado"])
        ws_ventas.cell(row=fila, column=8, value=v["agregado_at"])
        ws_ventas.cell(row=fila, column=9, value=v["motivo_cancelacion"] or "")
        if v["estado"] == "confirmado":
            total_confirmado += subtotal_cup
        fila += 1
    ws_ventas.cell(row=fila + 1, column=5, value="Total confirmado:").font = Font(name=FUENTE, bold=True)
    ws_ventas.cell(row=fila + 1, column=6, value=total_confirmado).number_format = FORMATO_CUP

    ws_pagos = wb.create_sheet("Pagos")
    _encabezado(ws_pagos, 1, ["ID pago", "Referencia", "Método", "Moneda", "Subtipo", "Monto", "Equivalente CUP", "Fecha"])
    fila = 2
    for p in d["pagos"]:
        ws_pagos.cell(row=fila, column=1, value=p["id"])
        ws_pagos.cell(row=fila, column=2, value=p["referencia"])
        ws_pagos.cell(row=fila, column=3, value=p["metodo"])
        ws_pagos.cell(row=fila, column=4, value=p["moneda"])
        ws_pagos.cell(row=fila, column=5, value=p["subtipo"] or "")
        ws_pagos.cell(row=fila, column=6, value=p["monto"] / 100)
        ws_pagos.cell(row=fila, column=7, value=p["monto_cup_equivalente"] / 100).number_format = FORMATO_CUP
        ws_pagos.cell(row=fila, column=8, value=p["registrado_at"])
        fila += 1

    ws_mov = wb.create_sheet("Movimientos de inventario")
    _hoja_movimientos(ws_mov, d["movimientos"])

    for prefijo, momento in (("Conteo apertura", "conteo_apertura"), ("Conteo cierre", "conteo_cierre")):
        ws_c = wb.create_sheet(prefijo)
        _encabezado(ws_c, 1, ["Insumo", "Sistema", "Contado", "Diferencia", "Nota", "Usuario", "Fecha"])
        fila = 2
        for c in d[momento]:
            ws_c.cell(row=fila, column=1, value=c["insumo_nombre"])
            ws_c.cell(row=fila, column=2, value=f'{c["cantidad_sistema"]} {c["unidad_medida"]}')
            ws_c.cell(row=fila, column=3, value=f'{c["cantidad_contada"]} {c["unidad_medida"]}')
            ws_c.cell(row=fila, column=4, value=c["diferencia"])
            ws_c.cell(row=fila, column=5, value=c["nota"] or "")
            ws_c.cell(row=fila, column=6, value=c["usuario_nombre"])
            ws_c.cell(row=fila, column=7, value=c["contado_at"])
            fila += 1
        if fila == 2:
            ws_c.cell(row=2, column=1, value="Sin conteo registrado")

    for ws in wb.worksheets:
        for col_idx in range(1, 10):
            ws.column_dimensions[chr(64 + col_idx)].width = 20
        for row in ws.iter_rows():
            for celda in row:
                if celda.font is None or not celda.font.bold:
                    celda.font = Font(name=FUENTE)

    return _xlsx_response(wb, f"cierre_{cierre['fecha']}.xlsx")


# ---------------------------------------------------------------------
# PDF -- pensado para imprimir o archivar, no para reprocesar datos.
# ---------------------------------------------------------------------

_ESTILOS = getSampleStyleSheet()
_ESTILO_TITULO = ParagraphStyle("TituloFP", parent=_ESTILOS["Title"], fontSize=16, spaceAfter=4)
_ESTILO_SUBTITULO = ParagraphStyle("SubtituloFP", parent=_ESTILOS["Normal"], fontSize=9, textColor=colors.grey)
_ESTILO_SECCION = ParagraphStyle("SeccionFP", parent=_ESTILOS["Heading2"], fontSize=12, spaceBefore=14, spaceAfter=6)


def _tabla_pdf(encabezados: list[str], filas: list[list[str]], anchos: list[float] | None = None) -> Table:
    data = [encabezados] + (filas if filas else [["—"] * len(encabezados)])
    t = Table(data, colWidths=anchos, repeatRows=1)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#B8871E")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return t


def _fmt_cup(centavos: int) -> str:
    return f"{centavos / 100:,.2f} CUP"


@router.get("/cierre/{cierre_id}.pdf")
def exportar_cierre_pdf(
    cierre_id: int, usuario_id: int, conn: sqlite3.Connection = Depends(get_conn)
) -> StreamingResponse:
    """Reporte imprimible de un día de caja: resumen, ventas, movimientos
    de inventario y ambos conteos. Pensado para archivar en papel o PDF,
    no para reimportar datos -- para eso está el .xlsx."""
    require_role(conn, usuario_id, *PUEDE_OPERAR_CAJA)
    d = _datos_dia_cierre(conn, cierre_id)
    cierre = d["cierre"]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    elementos = []

    elementos.append(Paragraph(f"FlowPos (Local) — Cierre de caja del {cierre['fecha']}", _ESTILO_TITULO))
    sub = f"Abrió {d['abierto_por_nombre']} · {cierre['hora_apertura']}"
    if cierre["hora_cierre"]:
        sub += f"  —  Cerró {d['cerrado_por_nombre'] or '—'} · {cierre['hora_cierre']}"
    else:
        sub += "  —  Caja en curso (todavía no cerrada)"
    elementos.append(Paragraph(sub, _ESTILO_SUBTITULO))
    elementos.append(Spacer(1, 10))

    if cierre["diferencia_cup"] is not None:
        diferencia = cierre["diferencia_cup"]
        etiqueta = "cuadrada" if diferencia == 0 else ("sobrante" if diferencia > 0 else "faltante")
        elementos.append(_tabla_pdf(
            ["Efectivo contado", "Efectivo esperado", "Diferencia"],
            [[_fmt_cup(cierre["efectivo_contado_cup"]), _fmt_cup(cierre["efectivo_esperado_cup"]),
              f"{_fmt_cup(abs(diferencia))} ({etiqueta})"]],
        ))
        elementos.append(Spacer(1, 6))

    elementos.append(Paragraph("Pagos por método", _ESTILO_SECCION))
    filas_resumen = [
        [f["metodo"], f["moneda"], f["subtipo"] or "—", str(f["cantidad"]), _fmt_cup(f["monto_cup_total"])]
        for f in d["resumen"]["desglose"]
    ]
    elementos.append(_tabla_pdf(["Método", "Moneda", "Subtipo", "Pagos", "Total"], filas_resumen))

    elementos.append(Paragraph(f"Ventas del día ({len(d['ventas'])})", _ESTILO_SECCION))
    filas_ventas = [
        [v["referencia"], v["producto"], str(v["cantidad"]), _fmt_cup(v["subtotal"]), v["estado"]]
        for v in d["ventas"]
    ]
    elementos.append(_tabla_pdf(["Cuenta", "Producto", "Cant.", "Subtotal", "Estado"], filas_ventas))

    elementos.append(Paragraph(f"Movimientos de inventario ({len(d['movimientos'])})", _ESTILO_SECCION))
    filas_mov = [
        [
            m["created_at"], m["insumo_nombre"], TIPO_MOVIMIENTO_LABEL.get(m["tipo"], m["tipo"]),
            f'{"+" if m["cantidad"] > 0 else ""}{m["cantidad"]} {m["unidad_medida"]}',
            m["usuario_nombre"], m["nota"] or "",
        ]
        for m in d["movimientos"]
    ]
    elementos.append(_tabla_pdf(
        ["Fecha", "Insumo", "Tipo", "Cant.", "Usuario", "Nota"], filas_mov,
        anchos=[3.2 * cm, 3.2 * cm, 3.6 * cm, 2.2 * cm, 2.5 * cm, 4 * cm],
    ))

    elementos.append(PageBreak())
    elementos.append(Paragraph("Conteo físico de apertura", _ESTILO_SECCION))
    filas_apertura = [
        [c["insumo_nombre"], f'{c["cantidad_sistema"]} {c["unidad_medida"]}',
         f'{c["cantidad_contada"]} {c["unidad_medida"]}', str(c["diferencia"]), c["nota"] or "—"]
        for c in d["conteo_apertura"]
    ]
    elementos.append(_tabla_pdf(["Insumo", "Sistema", "Contado", "Dif.", "Nota"], filas_apertura))

    elementos.append(Paragraph("Conteo físico de cierre", _ESTILO_SECCION))
    filas_cierre = [
        [c["insumo_nombre"], f'{c["cantidad_sistema"]} {c["unidad_medida"]}',
         f'{c["cantidad_contada"]} {c["unidad_medida"]}', str(c["diferencia"]), c["nota"] or "—"]
        for c in d["conteo_cierre"]
    ]
    elementos.append(_tabla_pdf(["Insumo", "Sistema", "Contado", "Dif.", "Nota"], filas_cierre))

    doc.build(elementos)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="cierre_{cierre["fecha"]}.pdf"'},
    )
